"""
Coach Slot Availability
=======================
Computes per-coach Total / Blocked / Booked / Available-for-booking / Open slots
from two exports:
  - slots file    : availability rules (one window per row; chopped into slots)
  - consumed file : bookings + blocks (one row per consumed slot-time)

Model:
  Available for booking = Total Slots - Blocked
  Open                  = Available for booking - Booked
  (each slot lands in exactly one bucket: Booked > Blocked > Open)

Matching is INTERVAL OVERLAP (strict): a defined slot [s,e] is consumed by a
booking/block [bs,be] iff  s < be AND e > bs  (touching edges do NOT overlap).
  - Booked  : slot start exactly equals a type-A (P/F/N) appointment start
  - Blocked : any other overlap (type-B block, or off-grid/custom appointment)
  - Open    : no overlap

Created 11-Jun-2026 IST.

Usage:
  python coach_availability.py --slots SLOTS.xlsx --consumed CONSUMED.xlsx --out OUT.xlsx \
      --start 2026-06-01 --end 2026-06-30 \
      [--status P,F,N] [--exclude-time-slot 60] [--exclude-coach Sahana] [--today 2026-06-11]
"""

import argparse, re, ast
from datetime import timedelta, datetime
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WEEKDAY = {"mon":0,"tue":1,"wed":2,"thu":3,"fri":4,"sat":5,"sun":6}


# ---------- helpers ----------
def read_any(path):
    return pd.read_excel(path) if str(path).lower().endswith((".xlsx",".xls")) else pd.read_csv(path)

def pick(df, base, which="first"):
    """Resolve a logical field to a real column. Duplicate columns from a join get
    pandas-suffixed (base, base.1). Slot/appointment times use the populated one;
    ids use the first occurrence."""
    pat = re.compile(rf"^{re.escape(base)}(\.\d+)?$")
    m = [c for c in df.columns if pat.match(c)]
    if not m:
        raise KeyError(f"'{base}' not found in {list(df.columns)}")
    return m[-1] if which == "last" else m[0]

def chop(start, end, step):
    """Cut a [start,end] HH:MM(:SS) window into step-minute slots -> [(start,end), ...]."""
    out = []
    t = datetime(2000,1,1,*map(int, str(start).split(":")[:2]))
    e = datetime(2000,1,1,*map(int, str(end).split(":")[:2]))
    while t < e:
        out.append((t.strftime("%H:%M"), (t+timedelta(minutes=int(step))).strftime("%H:%M")))
        t += timedelta(minutes=int(step))
    return out


def _hhmm(t):
    """Normalise a time to HH:MM (handles '18:45:00', '9:00')."""
    m = re.match(r"(\d{1,2}):(\d{2})", str(t)) if t is not None else None
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else None

def _min(hhmm):
    """HH:MM -> minutes since midnight."""
    h, m = str(hhmm).split(":")[:2]
    return int(h) * 60 + int(m)

def _reserved_windows(row):
    """reserved_slot_config.week1..week4 -> {week:int -> set((start,end))} in HH:MM.
    Handles both live (lists) and Excel (str-repr) forms; missing -> {}.
    Week of month = days 1-7 -> 1, 8-14 -> 2, 15-21 -> 3, 22+ -> 4."""
    out = {}
    for wk in (1, 2, 3, 4):
        val = row.get(f"reserved_slot_config.week{wk}")
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        if isinstance(val, str):
            try: val = ast.literal_eval(val)
            except Exception: continue
        if not isinstance(val, list):
            continue
        s = {(_hhmm(it.get("start_time")), _hhmm(it.get("end_time")))
             for it in val if isinstance(it, dict)}
        s = {w for w in s if w[0] and w[1]}
        if s: out[wk] = s
    return out


# ---------- slot grid ----------
def build_slots(slots, win_start, win_end, exclude_durations, exclude_coaches):
    df = slots.copy() if isinstance(slots, pd.DataFrame) else read_any(slots)   # path or DataFrame
    ia, idl = pick(df,"is_active","first"), pick(df,"is_deleted","first")
    df = df[(df[ia]=="Y") & (df[idl]=="N")].copy()
    if exclude_durations:
        df = df[~df["time_slot"].isin(exclude_durations)]
    df["sd"] = pd.to_datetime(df["start_date"], errors="coerce")
    df["ed"] = pd.to_datetime(df["end_date"], errors="coerce")
    cid = pick(df,"health_coach_id","first"); pst = pick(df,"start_time","first"); pet = pick(df,"end_time","first")

    # reserved windows are coach-level: a rule's reserved_slot_config can describe
    # times belonging to that coach's OTHER rules, so aggregate per coach per week.
    coach_res = {}                                              # cid -> {week -> [(start_min, end_min)]}
    for _, r in df.iterrows():
        rw = _reserved_windows(r)
        if not rw:
            continue
        dst = coach_res.setdefault(r[cid], {})
        for wk, wins in rw.items():
            dst.setdefault(wk, []).extend((_min(a), _min(b)) for a, b in wins)

    rows = []
    for _, r in df.iterrows():
        lo, hi = max(r["sd"], win_start), min(r["ed"], win_end)     # clip rule to window
        if pd.isna(lo) or pd.isna(hi) or lo > hi:
            continue
        slots = chop(r[pst], r[pet], r["time_slot"])
        days = {WEEKDAY[d.strip()] for d in str(r["days"]).split(",") if d.strip() in WEEKDAY}
        cres = coach_res.get(r[cid], {})
        d = lo
        while d <= hi:
            if d.weekday() in days:
                wins = cres.get(min(4, (d.day - 1)//7 + 1), ())     # reserved windows for this week-of-month
                for s_, e_ in slots:
                    sm, em = _min(s_), _min(e_)
                    reserved = any(a <= sm and em <= b for a, b in wins)   # slot fully inside a reserved window
                    rows.append((r[cid], d.date(), s_, e_,
                                 r.get("first_name",""), r.get("last_name",""), r.get("role",""),
                                 reserved))
            d += timedelta(days=1)

    inst = pd.DataFrame(rows, columns=["coach","date","t","tend","first","last","role","reserved"]) \
             .drop_duplicates(["coach","date","t"])
    inst["name"] = (inst["first"].fillna("")+" "+inst["last"].fillna("")).str.strip()
    if exclude_coaches:
        pat = "|".join(re.escape(c) for c in exclude_coaches)
        inst = inst[~inst["name"].str.contains(pat, case=False, na=False)]
    return inst


# ---------- consumed ----------
def load_consumed(consumed, statuses):
    df = consumed.copy() if isinstance(consumed, pd.DataFrame) else read_any(consumed)   # path or DataFrame
    cid = pick(df,"health_coach_id","first")
    df["cid"] = df[cid]
    df["s"] = pd.to_datetime(df[pick(df,"appointment_start_time","last")], errors="coerce")
    df["e"] = pd.to_datetime(df[pick(df,"appointment_end_time","last")], errors="coerce")
    df["date"] = df["s"].dt.date
    df["st"] = df["s"].dt.strftime("%H:%M")

    appt = df[(df["status"].isin(statuses)) & (df["type"]=="A")]              # patient appts
    cons = df[(df["status"].isin(statuses)) | (df["type"]=="B")]             # everything consuming
    blkw = df[(df["type"]=="B") & (df["block_whole_day"]=="Y")]

    cons_by = defaultdict(list)
    for _, r in cons.iterrows():
        if pd.notna(r["s"]) and pd.notna(r["e"]):
            cons_by[(r["cid"], r["date"])].append((r["s"].to_pydatetime(), r["e"].to_pydatetime()))
    appt_exact = set(zip(appt["cid"], appt["date"], appt["st"]))
    bwd_keys   = set(zip(blkw["cid"], blkw["date"], blkw["st"]))
    return appt_exact, cons_by, bwd_keys


# ---------- classify ----------
def classify(inst, appt_exact, cons_by, bwd_keys):
    def bucket(row):
        c, d = row["coach"], row["date"]
        if (c, d, row["t"]) in appt_exact:
            return "booked"
        ss = datetime.combine(d, datetime.strptime(row["t"], "%H:%M").time())
        se = datetime.combine(d, datetime.strptime(row["tend"], "%H:%M").time())
        for bs, be in cons_by.get((c, d), []):
            if ss < be and se > bs:        # strict overlap; touching edges excluded
                return "blocked"
        return "open"
    inst = inst.copy()
    inst["bucket"]  = inst.apply(bucket, axis=1)
    inst["booked"]  = inst["bucket"]=="booked"
    inst["blocked"] = inst["bucket"]=="blocked"
    inst["open"]    = inst["bucket"]=="open"
    inst["blk_wd"]  = [(b and (c,d,t) in bwd_keys)
                       for b,c,d,t in zip(inst["blocked"],inst["coach"],inst["date"],inst["t"])]
    return inst


# ---------- summarise ----------
def summarise(inst, today):
    inst = inst.copy()
    inst["role"] = inst["role"].fillna("")          # 11-Jun-2026: API has no role; avoid NaN groupby key dropping rows
    inst["dt"] = pd.to_datetime(inst["date"].astype(str)+" "+inst["t"])
    up = inst[inst["open"] & (inst["dt"] >= today)]
    o = up.groupby("name").agg(open_up=("dt","size"), next_open=("dt","min"), last_open=("dt","max")).reset_index()
    g = inst.groupby(["name","role"], dropna=False).agg(
        total=("t","size"), booked=("booked","sum"),
        blocked=("blocked","sum"), blk_wd=("blk_wd","sum"), open=("open","sum")).reset_index()
    g = g.merge(o, on="name", how="left")
    g["open_up"] = g["open_up"].fillna(0).astype(int)
    g["avail_for_booking"] = g["total"] - g["blocked"]
    return g.sort_values("open", ascending=False)


def next_open_slots(slots, consumed, n0, n1, statuses, excl_dur=frozenset(), excl_co=()):
    """Enumerate OPEN slots in the window [n0, n1] by actual date.
    Returns df: name, role, date, day, start, end (sorted by coach, date, time)."""
    inst = build_slots(slots, n0, n1, set(excl_dur), list(excl_co))
    appt_exact, cons_by, bwd = load_consumed(consumed, statuses)
    inst = classify(inst, appt_exact, cons_by, bwd)
    op = inst[inst["open"]].copy()
    op["role"] = op["role"].fillna("")
    if op.empty:
        return pd.DataFrame(columns=["name","role","date","day","start","end","reserved"])
    op["date"] = pd.to_datetime(op["date"])
    op["day"] = op["date"].dt.strftime("%a")
    op = op.sort_values(["name","date","t"]).rename(columns={"t":"start","tend":"end"})
    return op[["name","role","date","day","start","end","reserved"]]


def calculate(slots, consumed, w0, w1, today, statuses, excl_dur=frozenset(), excl_co=()):
    """Run the full calc for window [w0,w1].
    Returns (g, inst, open7, (n0,n1)). g carries per-coach 'open_7d' =
    count of open slots in the 7 days AFTER today."""
    inst = build_slots(slots, w0, w1, set(excl_dur), list(excl_co))
    appt_exact, cons_by, bwd = load_consumed(consumed, statuses)
    inst = classify(inst, appt_exact, cons_by, bwd)
    g = summarise(inst, today)
    n0, n1 = today + pd.Timedelta(days=1), today + pd.Timedelta(days=7)
    open7 = next_open_slots(slots, consumed, n0, n1, statuses, excl_dur, excl_co)
    cnt = open7.groupby("name").size() if len(open7) else pd.Series(dtype="int64")
    g["open_7d"] = g["name"].map(cnt).fillna(0).astype(int)
    return g, inst, open7, (n0, n1)


# ---------- excel ----------
F="Arial"
HF=PatternFill("solid",start_color="1F4E78"); HFONT=Font(name=F,bold=True,color="FFFFFF",size=11)
TF=PatternFill("solid",start_color="DDEBF7"); BF=PatternFill("solid",start_color="FCE4D6"); GF=PatternFill("solid",start_color="E2EFDA")
TH=Side(style="thin",color="BFBFBF"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)

def write_excel(g, inst, out, meta, open7=None):
    wb = Workbook(); ws = wb.active; ws.title = "Coach Availability"
    H=["Coach","Role","Total Slots","Blocked","of which whole-day","Booked (appointments)",
       "Available for booking","Open","Open (upcoming)","Next Open","Last Open"]
    ws.append(H)
    for c in ws[1]: c.fill,c.font,c.border=HF,HFONT,BD; c.alignment=Alignment(horizontal="center",wrap_text=True)
    r=2
    for _,row in g.iterrows():
        ws.cell(r,1,row["name"]); ws.cell(r,2,row["role"]); ws.cell(r,3,int(row["total"]))
        ws.cell(r,4,int(row["blocked"])); ws.cell(r,5,int(row["blk_wd"])); ws.cell(r,6,int(row["booked"]))
        ws.cell(r,7,f"=C{r}-D{r}"); ws.cell(r,8,f"=G{r}-F{r}"); ws.cell(r,9,int(row["open_up"]))
        ws.cell(r,10,row["next_open"].strftime("%d-%b %H:%M") if pd.notna(row["next_open"]) else "-")
        ws.cell(r,11,row["last_open"].strftime("%d-%b %H:%M") if pd.notna(row["last_open"]) else "-")
        for col in range(1,12):
            cell=ws.cell(r,col); cell.font,cell.border=Font(name=F,size=11),BD
            if col>=3: cell.alignment=Alignment(horizontal="center")
            if col in (4,5): cell.fill=BF
            if col in (7,8): cell.fill=GF
        r+=1
    ws.cell(r,1,"TOTAL").font=Font(name=F,bold=True,size=11)
    for col,L in [(3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I")]: ws.cell(r,col,f"=SUM({L}2:{L}{r-1})")
    for col in range(1,12):
        cell=ws.cell(r,col); cell.fill,cell.font,cell.border=TF,Font(name=F,bold=True,size=11),BD
        if col>=3: cell.alignment=Alignment(horizontal="center")
    for col,w in zip("ABCDEFGHIJK",[22,16,11,9,15,18,18,8,14,13,13]): ws.column_dimensions[col].width=w
    ws.freeze_panes="C2"; ws.row_dimensions[1].height=30

    ws2 = wb.create_sheet("Slot Detail (audit)")
    det = inst.copy()
    det["status"] = det.apply(lambda r:"Blocked (whole-day)" if r["blk_wd"] else
                              ("Blocked" if r["blocked"] else ("Booked" if r["booked"] else "Open")), axis=1)
    det = det[["name","role","date","t","tend","status","reserved"]].sort_values(["name","date","t"])
    ws2.append(["Coach","Role","Date","Slot Start","Slot End","Status","Reserved"])
    for c in ws2[1]: c.fill,c.font,c.border=HF,HFONT,BD; c.alignment=Alignment(horizontal="center")
    for _,row in det.iterrows(): ws2.append([row["name"],row["role"],str(row["date"]),row["t"],row["tend"],row["status"],
                                             "Yes" if row["reserved"] else ""])
    for col,w in zip("ABCDEFG",[22,16,12,10,10,18,9]): ws2.column_dimensions[col].width=w
    ws2.freeze_panes="A2"

    if open7 is not None:
        ws4 = wb.create_sheet("Open Slots (next 7d)")
        o = open7.copy()
        if len(o):
            o["date"] = pd.to_datetime(o["date"])
            o = o.sort_values(["name", "date", "start"])
            o["label"] = o.apply(lambda r: f"{r['date']:%d-%b} {r['start']}-{r['end']}"
                                            + (" (R)" if r.get("reserved") else ""), axis=1)  # date + time (+R if reserved)
            grp = o.groupby("name", sort=True)["label"].apply(list)
        else:
            grp = pd.Series(dtype=object)
        maxn = max((len(v) for v in grp), default=0)
        ws4.append(["Coach"] + [f"Slot {i}" for i in range(1, maxn + 1)])
        for c in ws4[1]: c.fill,c.font,c.border=HF,HFONT,BD; c.alignment=Alignment(horizontal="center")
        for name, labels in grp.items():
            ws4.append([name] + labels)
        ws4.column_dimensions["A"].width = 22
        for i in range(2, maxn + 2):
            ws4.column_dimensions[get_column_letter(i)].width = 16
        ws4.freeze_panes = "B2"

    ws3 = wb.create_sheet("Method & Notes")
    for k,v in meta.items(): ws3.append([k,v])
    for row in ws3.iter_rows():
        for c in row: c.font=Font(name=F,size=11,bold=(c.column==1))
    ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=72
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slots", required=True)
    ap.add_argument("--consumed", required=True)
    ap.add_argument("--out", default="coach_availability.xlsx")
    ap.add_argument("--start", required=True, help="window start YYYY-MM-DD")
    ap.add_argument("--end", required=True, help="window end YYYY-MM-DD")
    ap.add_argument("--today", help="open/upcoming cutoff YYYY-MM-DD (default: now)")
    ap.add_argument("--status", default="P,F,N", help="occupied statuses (default P,F,N)")
    ap.add_argument("--exclude-time-slot", default="", help="durations to drop, e.g. 60")
    ap.add_argument("--exclude-coach", default="", help="coach name substrings to drop, comma-separated")
    a = ap.parse_args()

    statuses = {s.strip() for s in a.status.split(",") if s.strip()}
    excl_dur = {int(x) for x in a.exclude_time_slot.split(",") if x.strip()}
    excl_co  = [c.strip() for c in a.exclude_coach.split(",") if c.strip()]
    w0, w1 = pd.Timestamp(a.start), pd.Timestamp(a.end)
    today = pd.Timestamp(a.today) if a.today else pd.Timestamp.now().normalize()

    g, inst, open7, (n0, n1) = calculate(a.slots, a.consumed, w0, w1, today, statuses, excl_dur, excl_co)

    print(g[["name","role","total","blocked","booked","avail_for_booking","open","open_7d"]].to_string(index=False))
    tot = dict(total=int(g["total"].sum()), blocked=int(g["blocked"].sum()),
               booked=int(g["booked"].sum()), avail=int(g["avail_for_booking"].sum()), open=int(g["open"].sum()))
    print("\nTOTALS:", tot)

    meta = {
        "Window": f"{w0.date()} to {w1.date()}",
        "Occupied statuses": ", ".join(sorted(statuses)),
        "Excluded durations": ", ".join(map(str, sorted(excl_dur))) or "none",
        "Excluded coaches": ", ".join(excl_co) or "none",
        "Matching": "interval overlap (strict); touching edges excluded",
        "Booked": "slot start = exact type-A appointment start",
        "Blocked": "slot overlaps any consumed interval (block or off-grid/custom appt)",
        "Available for booking": "Total - Blocked",
        "Open": "Available for booking - Booked",
        "Open (upcoming)": f"open slots dated {today.date()} onward",
        "Open (next 7d)": f"open slots {n0.date()} to {n1.date()} (after today)",
        "TOTALS": str(tot),
    }
    write_excel(g, inst, a.out, meta, open7=open7)
    print("saved", a.out)


if __name__ == "__main__":
    main()